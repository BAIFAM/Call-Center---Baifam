
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from drf_spectacular.utils import extend_schema, OpenApiParameter, OpenApiResponse, OpenApiTypes
from django.shortcuts import get_object_or_404

from users.models import Profile
from .models import Agent, CallGroup, CallGroupContact, CallGroupAgent, Contact, Call, ContactProduct
from institution.models import Institution, Product
from .serializers import AgentSerializer, BulkContactSerializer, CallGroupContactSerializer, CallGroupSerializer, CallGroupAgentSerializer, CallSerializer, ContactProductSerializer, ContactSerializer
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
import pandas as pd
from django.http import HttpResponse
import io
import uuid as uuid_module
import logging
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)
@extend_schema(tags=["CallGroup"])
class CallGroupListCreateView(APIView):

    @extend_schema(
        summary="List all call groups for a specific institution",
        parameters=[
            OpenApiParameter(name="institution_id", required=True, type=int, location=OpenApiParameter.PATH),
        ],
        responses={200: CallGroupSerializer(many=True)}
    )
    def get(self, request, institution_id):
        institution = get_object_or_404(Institution, id=institution_id)
        groups = CallGroup.objects.filter(institution=institution)
        serializer = CallGroupSerializer(groups, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @extend_schema(
        summary="Create a new call group for an institution",
        request=CallGroupSerializer,
        responses={201: CallGroupSerializer}
    )
    def post(self, request, institution_id):
        institution = get_object_or_404(Institution, id=institution_id)
        data = request.data.copy()
        data['institution'] = str(institution.id)
        serializer = CallGroupSerializer(data=data)
        if serializer.is_valid():
            serializer.save(created_by=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@extend_schema(tags=["CallGroup"])
class CallGroupDetailView(APIView):

    def get_object(self, uuid):
        return get_object_or_404(CallGroup, uuid=uuid)

    @extend_schema(
        summary="Retrieve a call group by UUID",
        responses={200: CallGroupSerializer}
    )
    def get(self, request, uuid):
        group = self.get_object(uuid)
        serializer = CallGroupSerializer(group)
        return Response(serializer.data)

    @extend_schema(
        summary="Partially update a call group by UUID",
        request=CallGroupSerializer,
        responses={200: CallGroupSerializer}
    )
    def patch(self, request, uuid):
        group = self.get_object(uuid)
        serializer = CallGroupSerializer(group, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        summary="Delete a call group by UUID",
        responses={204: OpenApiResponse(description="Deleted successfully")}
    )
    def delete(self, request, uuid):
        group = self.get_object(uuid)
        group.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
@extend_schema(tags=["CallGroupAgent"])
class CallGroupAgentListCreateView(APIView):

    @extend_schema(
        summary="List users assigned to call groups of a specific institution",
        parameters=[
            OpenApiParameter(name="institution_id", required=True, type=int, location=OpenApiParameter.PATH),
        ],
        responses={200: CallGroupAgentSerializer(many=True)}
    )
    def get(self, request, institution_id):
        institution = get_object_or_404(Institution, id=institution_id)
        users = CallGroupAgent.objects.filter(call_group__institution=institution)
        serializer = CallGroupAgentSerializer(users, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @extend_schema(
        summary="Assign a user to a call group (within specified institution)",
        request=CallGroupAgentSerializer,
        responses={201: CallGroupAgentSerializer}
    )
    def post(self, request, institution_id):
        institution = get_object_or_404(Institution, id=institution_id)
        data = request.data.copy()

        call_group_uuid = data.get('call_group')
        call_group = get_object_or_404(CallGroup, uuid=call_group_uuid, institution=institution)

        serializer = CallGroupAgentSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@extend_schema(tags=["CallGroupAgent"])
class CallGroupAgentDetailView(APIView):

    def get_object(self, uuid):
        return get_object_or_404(CallGroupAgent, uuid=uuid)

    @extend_schema(
        summary="Retrieve a CallGroupAgent by UUID",
        responses={200: CallGroupAgentSerializer}
    )
    def get(self, request, uuid):
        item = self.get_object(uuid)
        serializer = CallGroupAgentSerializer(item)
        return Response(serializer.data)

    @extend_schema(
        summary="Partially update a CallGroupAgent by UUID",
        request=CallGroupAgentSerializer,
        responses={200: CallGroupAgentSerializer}
    )
    def patch(self, request, uuid):
        item = self.get_object(uuid)
        serializer = CallGroupAgentSerializer(item, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        summary="Delete a CallGroupAgent by UUID",
        responses={204: OpenApiResponse(description="Deleted successfully")}
    )
    def delete(self, request, uuid):
        item = self.get_object(uuid)
        item.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)    

@extend_schema(tags=["CallGroup"])
class UserCallGroupsListView(APIView):
    @extend_schema(
        summary="List call groups assigned to the authenticated user",
        responses={200: CallGroupSerializer(many=True)}
    )
    def get(self, request, institution_id):
        user = request.user
        groups = CallGroup.objects.filter(users__user=user, institution__id=institution_id, users__status="active").distinct()
        serializer = CallGroupSerializer(groups, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)



@extend_schema(tags=["Contact"])
class ContactListCreateView(APIView):

    @extend_schema(
        summary="List Contacts for a specific institution",
        parameters=[
            OpenApiParameter(name="institution_id", required=True, type=int, location=OpenApiParameter.PATH),
        ],
        responses={200: ContactSerializer(many=True)}
    )
    def get(self, request, institution_id):
        institution = get_object_or_404(Institution, id=institution_id)
        contacts = Contact.objects.filter(institution=institution).prefetch_related('contact_products__product__institution')
        serializer = ContactSerializer(contacts, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @extend_schema(
        summary="Create a Contact (within specified institution)",
        request=ContactSerializer,
        responses={201: ContactSerializer}
    )
    def post(self, request, institution_id):
        institution = get_object_or_404(Institution, id=institution_id)
        data = request.data.copy()
        data['institution'] = institution.id

        serializer = ContactSerializer(data=data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@extend_schema(tags=["Contact"])
class ContactTemplateDownloadView(APIView):
    @extend_schema(
        summary="Download Excel template for bulk contact creation",
        parameters=[
            OpenApiParameter(name="product_uuid", required=True, type=str, location=OpenApiParameter.PATH),
        ],
        responses={200: OpenApiResponse(description="Excel template file")}
    )
    def get(self, request, product_uuid):
        product = get_object_or_404(Product, uuid=product_uuid)
        
        try:
            # Create template with empty rows
            template_data = {
                'name': [''] * 10,
                'phone_number': [''] * 10,
                'country': [''] * 10,
                'country_code': [''] * 10,
                'status': [''] * 10,
                'remarks': [''] * 10,
            }
            
            # Create Excel file in memory
            output = io.BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Main template sheet
                df = pd.DataFrame(template_data)
                df.to_excel(writer, sheet_name='Contacts', index=False)
                
                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['Contacts']
                
                # Set column widths
                column_widths = {
                    'A': 25,  # name
                    'B': 20,  # phone_number
                    'C': 15,  # country
                    'D': 15,  # country_code
                    'E': 15,  # status
                    'F': 30,  # remarks
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # Header styling
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                center_alignment = Alignment(horizontal="center", vertical="center")
                
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                
                # Instructions sheet
                instructions_df = pd.DataFrame({
                    'Field': ['name', 'phone_number', 'country', 'country_code', 'status', 'remarks'],
                    'Description': [
                        'Full name of the contact',
                        'Phone number with country code',
                        'Country name (optional)',
                        'Country calling code (optional)',
                        'Contact status (optional, default: new)',
                        'Additional remarks (optional)'
                    ],
                    'Required': ['Yes', 'Yes', 'No', 'No', 'No', 'No'],
                    'Examples': [
                        'John Doe',
                        '+256701234567',
                        'Uganda',
                        '+256',
                        'new, assigned, attended_to, archived, flagged, ready_to_export, exported',
                        'Met at conference'
                    ]
                })
                instructions_df.to_excel(writer, sheet_name='Instructions', index=False)
                
                instructions_sheet = writer.sheets['Instructions']
                for col in ['A', 'B', 'C', 'D']:
                    instructions_sheet.column_dimensions[col].width = 30
                
                for col in range(1, len(instructions_df.columns) + 1):
                    cell = instructions_sheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
                
                # Product Info sheet
                product_info_df = pd.DataFrame({
                    'Product Information': [
                        'Selected Product',
                        'Product Name',
                        'Institution',
                        'Upload Instructions',
                        '',
                        '1. Fill in the Contacts sheet',
                        '2. Save the file',
                        '3. Upload using the bulk upload feature',
                        '4. Contacts will be linked to the selected product'
                    ],
                    'Details': [
                        '',
                        product.name,
                        product.institution.institution_name,
                        '',
                        '',
                        '',
                        '',
                        '',
                        ''
                    ]
                })
                product_info_df.to_excel(writer, sheet_name='Product_Info', index=False)
                
                product_info_sheet = writer.sheets['Product_Info']
                product_info_sheet.column_dimensions['A'].width = 30
                product_info_sheet.column_dimensions['B'].width = 30
                
                for col in range(1, len(product_info_df.columns) + 1):
                    cell = product_info_sheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
            
            output.seek(0)
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                content=output.read()
            )
            response['Content-Disposition'] = f'attachment; filename="contacts_template_{product.name.replace(' ', '_')}.xlsx"'
            return response
            
        except Exception as e:
            return Response({'error': f'Failed to generate template: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@extend_schema(tags=["Contact"])
class ContactBulkUploadView(APIView):
    parser_classes = [MultiPartParser]

    @extend_schema(
        summary="Bulk upload contacts for a specific product",
        parameters=[
            OpenApiParameter(name="product_uuid", required=True, type=str, location=OpenApiParameter.PATH),
        ],
        request={
            'multipart/form-data': {
                'type': 'object',
                'properties': {
                    'file': {
                        'type': 'string',
                        'format': 'binary',
                        'description': 'Excel or CSV file with contact data'
                    }
                }
            }
        },
        responses={
            201: OpenApiResponse(description="Contacts created successfully"),
            400: OpenApiResponse(description="Validation errors")
        }
    )
    def post(self, request, product_uuid):
        product = get_object_or_404(Product, uuid=product_uuid)
        
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        
        # Validate file type
        if not file.name.endswith(('.xlsx', '.xls', '.csv')):
            return Response(
                {'error': 'Invalid file type. Please upload a CSV or Excel file (.csv, .xlsx, .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Read file
            if file.name.endswith('.csv'):
                df = pd.read_csv(file)
                df = df[~df.iloc[:, 0].astype(str).str.startswith('#')]
            else:
                df = pd.read_excel(file, sheet_name='Contacts')
            
            # Remove empty rows
            df = df.dropna(how='all')
            
            # Validate required columns
            required_columns = ['name', 'phone_number']
            missing_columns = [col for col in required_columns if col not in df.columns]
            if missing_columns:
                return Response(
                    {'error': f'Missing required columns: {", ".join(missing_columns)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            created_contacts = []
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Skip empty rows
                    if pd.isna(row['name']) or str(row['name']).strip() == '':
                        continue
                    
                    if pd.isna(row['phone_number']) or str(row['phone_number']).strip() == '':
                        errors.append(f'Row {index + 2}: Phone number is required')
                        continue
                    
                    # Prepare contact data
                    contact_data = {
                        'name': str(row['name']).strip(),
                        'phone_number': str(row['phone_number']).strip(),
                        'country': str(row.get('country', '')).strip() if pd.notna(row.get('country')) else '',
                        'country_code': str(row.get('country_code', '')).strip() if pd.notna(row.get('country_code')) else '',
                        'status': str(row.get('status', 'new')).strip() if pd.notna(row.get('status')) else 'new',
                        'remarks': str(row.get('remarks', '')).strip() if pd.notna(row.get('remarks')) else '',
                        'product': str(product.uuid)
                    }
                    
                    # Create contact using BulkContactSerializer
                    serializer = BulkContactSerializer(data=contact_data, context={'request': request})
                    if serializer.is_valid():
                        contact = serializer.save()
                        created_contacts.append({
                            'uuid': str(contact.uuid),
                            'name': contact.name,
                            'phone_number': contact.phone_number,
                            'product_name': product.name
                        })
                    else:
                        errors.append(f'Row {index + 2}: {serializer.errors}')
                        
                except Exception as e:
                    errors.append(f'Row {index + 2}: {str(e)}')
            
            response_data = {
                'created_count': len(created_contacts),
                'error_count': len(errors),
                'product_name': product.name,
                'institution': product.institution.institution_name,
                'created_contacts': created_contacts
            }
            
            if errors:
                response_data['errors'] = errors
            
            response_status = status.HTTP_201_CREATED if created_contacts else status.HTTP_400_BAD_REQUEST
            return Response(response_data, status=response_status)
            
        except Exception as e:
            return Response(
                {'error': f'Error processing file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )


@extend_schema(tags=["Contact"])
class ContactDetailView(APIView):

    def get_object(self, uuid):
        return get_object_or_404(Contact, uuid=uuid)

    @extend_schema(
        summary="Retrieve a Contact by UUID",
        parameters=[
            OpenApiParameter(name="uuid", required=True, type=str, location=OpenApiParameter.PATH),
        ],
        responses={200: ContactSerializer}
    )
    def get(self, request, uuid):
        item = self.get_object(uuid)
        serializer = ContactSerializer(item)
        return Response(serializer.data)

    @extend_schema(
        summary="Partially update a Contact by UUID",
        request=ContactSerializer,
        responses={200: ContactSerializer}
    )
    def patch(self, request, uuid):
        item = self.get_object(uuid)
        serializer = ContactSerializer(item, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        summary="Delete a Contact by UUID",
        responses={204: OpenApiResponse(description="Deleted successfully")}
    )
    def delete(self, request, uuid):
        item = self.get_object(uuid)
        item.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


@extend_schema(tags=["CallGroupContact"])    
class ContactsByCallGroupContactListCreateView(APIView):
    @extend_schema(
        summary="List contacts assigned to a specific call group",
        parameters=[
            OpenApiParameter(name="call_group_uuid", required=True, type=int, location=OpenApiParameter.PATH),
        ],
        responses={200: CallGroupContactSerializer(many=True)}
    )
    def get(self, request, call_group_uuid):
        group = get_object_or_404(CallGroup, uuid=call_group_uuid)
        contacts = CallGroupContact.objects.filter(call_group=group)
        serializer = CallGroupContactSerializer(contacts, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

class ContactCallsListView(APIView):
    @extend_schema(
        summary="List all calls made to a specific contact",
        parameters=[
            OpenApiParameter(name="contact_uuid", required=True, type=str, location=OpenApiParameter.PATH),
        ],
        responses={200: CallSerializer(many=True)}
    )
    def get(self, request, contact_uuid):
        contact = get_object_or_404(Contact, uuid=contact_uuid)
        calls = Call.objects.filter(contact=contact)
        serializer = CallSerializer(calls, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

@extend_schema(tags=["CallGroupContact"])
class CallGroupContactListCreateView(APIView):

    @extend_schema(
        summary="List contacts assigned to call groups of a specific institution",
        parameters=[
            OpenApiParameter(name="institution_id", required=True, type=int, location=OpenApiParameter.PATH),
        ],
        responses={200: CallGroupContactSerializer(many=True)}
    )
    def get(self, request, institution_id):
        institution = get_object_or_404(Institution, id=institution_id)
        contacts = CallGroupContact.objects.filter(call_group__institution=institution)
        serializer = CallGroupContactSerializer(contacts, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @extend_schema(
        summary="Assign a contact to a call group (within specified institution)",
        request=CallGroupContactSerializer,
        responses={201: CallGroupContactSerializer}
    )
    def post(self, request, institution_id):
        institution = get_object_or_404(Institution, id=institution_id)
        data = request.data.copy()

        # Validate that the selected call_group belongs to the institution
        call_group_uuid = data.get('call_group')
        call_group = get_object_or_404(CallGroup, uuid=call_group_uuid, institution=institution)

        serializer = CallGroupContactSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@extend_schema(tags=["CallGroupContact"])
class CallGroupContactDetailView(APIView):

    def get_object(self, uuid):
        return get_object_or_404(CallGroupContact, uuid=uuid)

    @extend_schema(
        summary="Retrieve a CallGroupContact by UUID",
        responses={200: CallGroupContactSerializer}
    )
    def get(self, request, uuid):
        item = self.get_object(uuid)
        serializer = CallGroupContactSerializer(item)
        return Response(serializer.data)

    @extend_schema(
        summary="Partially update a CallGroupContact by UUID",
        request=CallGroupContactSerializer,
        responses={200: CallGroupContactSerializer}
    )
    def patch(self, request, uuid):
        item = self.get_object(uuid)
        serializer = CallGroupContactSerializer(item, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        summary="Delete a CallGroupContact by UUID",
        responses={204: OpenApiResponse(description="Deleted successfully")}
    )
    def delete(self, request, uuid):
        item = self.get_object(uuid)
        item.delete()
        return Response(status=status.HTTP_204_NO_CONTENT) 
    
@extend_schema(tags=['Calls'])
class CallListCreateAPIView(APIView):
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    @extend_schema(
        summary='List all calls for a specific institution',
        parameters=[
            OpenApiParameter(name='institution_id', type=int, location=OpenApiParameter.PATH)
        ],
        responses={200: CallSerializer(many=True)}
    )
    def get(self, request, institution_id):
        calls = Call.objects.filter(contact__product__institution__id=institution_id)
        serializer = CallSerializer(calls, many=True, context={'request': request})
        return Response(serializer.data, status=status.HTTP_200_OK)

    @extend_schema(
        summary='Create a new call under an institution',
        request=CallSerializer,
        responses={201: CallSerializer}
    )
    def post(self, request, institution_id):
        contact_uuid = request.data.get('contact')
        if not contact_uuid:
            return Response(
                {"detail": "ContactProduct UUID is required."},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Retrieve ContactProduct instance
        contact_product = get_object_or_404(ContactProduct, uuid=contact_uuid)
        # Ensure contact_product's product belongs to the specified institution
        if contact_product.product.institution.id != institution_id:
            return Response(
                {'detail': 'ContactProduct does not belong to this institution.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Create a mutable copy of the data
        data = request.data.copy()

        # Add all files from request.FILES to data
        for field_name, file in request.FILES.items():
            data[field_name] = file

        print("\n\n\n Creating call with data : ", data)
        # Pass request context to serializer for dynamic file field creation
        serializer = CallSerializer(data=data, context={'request': request})
        if serializer.is_valid():
            # Validate made_by (optional, depending on your requirements)
            try:
                callgroup_agent = request.user.callgroupagent
                if callgroup_agent.institution.id != institution_id:
                    return Response(
                        {'detail': 'User does not belong to this institution.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            except AttributeError:
                return Response(
                    {'detail': 'User is not a valid CallgroupAgent.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            call = serializer.save(made_by=callgroup_agent)
            call.refresh_from_db()

            # Create a new serializer instance with the refreshed call for the response
            response_serializer = CallSerializer(call, context={'request': request})
            response_data = response_serializer.data

            return Response(response_data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    
@extend_schema(tags=['Calls'])
class CallDetailAPIView(APIView):

    def get_object(self, uuid):
        return get_object_or_404(Call, uuid=uuid)

    @extend_schema(
        summary='Retrieve a call by UUID',
        responses={200: CallSerializer}
    )
    def get(self, request, uuid):
        call = self.get_object(uuid)
        serializer = CallSerializer(call)
        return Response(serializer.data)

    @extend_schema(
        summary='Partially update a call',
        request=CallSerializer,
        responses={200: CallSerializer}
    )
    def patch(self, request, uuid):
        call = self.get_object(uuid)
        serializer = CallSerializer(call, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        summary='Delete a call',
        responses={204: OpenApiResponse(description='Deleted successfully')}
    )
    def delete(self, request, uuid):
        call = self.get_object(uuid)
        call.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

@extend_schema(tags=["ContactProduct"])
class ContactProductListCreateView(APIView):

    @extend_schema(
        summary="List ContactProduct associations for a specific institution",
        parameters=[
            OpenApiParameter(name="institution_id", required=True, type=int, location=OpenApiParameter.PATH),
        ],
        responses={200: ContactProductSerializer(many=True)}
    )
    def get(self, request, institution_id):
        institution = get_object_or_404(Institution, id=institution_id)
        contact_products = ContactProduct.objects.filter(product__institution=institution).select_related('contact', 'product__institution', 'created_by')
        serializer = ContactProductSerializer(contact_products, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @extend_schema(
        summary="Create a ContactProduct association (within specified institution)",
        request=ContactProductSerializer,
        responses={201: ContactProductSerializer}
    )
    def post(self, request, institution_id):
        institution = get_object_or_404(Institution, id=institution_id)
        data = request.data.copy()

        contact_id = data.get('contact')
        product_id = data.get('product')
        # Use uuid for both Contact and Product
        contact = get_object_or_404(Contact, uuid=contact_id)
        product = get_object_or_404(Product, uuid=product_id, institution=institution)

        serializer = ContactProductSerializer(data=data, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@extend_schema(tags=["ContactProduct"])
class ContactProductDetailView(APIView):

    def get_object(self, uuid):
        return get_object_or_404(ContactProduct, uuid=uuid)

    @extend_schema(
        summary="Retrieve a ContactProduct by UUID",
        parameters=[
            OpenApiParameter(name="uuid", required=True, type=OpenApiTypes.UUID, location=OpenApiParameter.PATH),
        ],
        responses={200: ContactProductSerializer}
    )
    def get(self, request, uuid):
        item = self.get_object(uuid)
        serializer = ContactProductSerializer(item)
        return Response(serializer.data)

    @extend_schema(
        summary="Partially update a ContactProduct by UUID",
        request=ContactProductSerializer,
        responses={200: ContactProductSerializer}
    )
    def patch(self, request, uuid):
        item = self.get_object(uuid)
        serializer = ContactProductSerializer(item, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        summary="Delete a ContactProduct by UUID",
        responses={204: OpenApiResponse(description="Deleted successfully")}
    )
    def delete(self, request, uuid):
        item = self.get_object(uuid)
        item.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

@extend_schema(tags=["Agent"])
class AgentListCreateView(APIView):

    @extend_schema(
        summary="List Agents for a specific institution",
        parameters=[
            OpenApiParameter(name="institution_id", required=True, type=int, location=OpenApiParameter.PATH),
        ],
        responses={200: AgentSerializer(many=True)}
    )
    def get(self, request, institution_id):
        institution = get_object_or_404(Institution, id=institution_id)
        agents = Agent.objects.filter(user__institution=institution).select_related('user__user', 'user__institution')
        serializer = AgentSerializer(agents, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    @extend_schema(
        summary="Create an Agent (within specified institution)",
        request=AgentSerializer,
        responses={201: AgentSerializer}
    )
    def post(self, request, institution_id):
        institution = get_object_or_404(Institution, id=institution_id)
        data = request.data.copy()

        user_id = data.get('user')
        user = get_object_or_404(Profile, id=user_id, institution=institution)

        serializer = AgentSerializer(data=data, context={'request': request, 'view': {'kwargs': {'institution_id': institution_id}}})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@extend_schema(tags=["Agent"])
class AgentDetailView(APIView):

    def get_object(self, uuid):
        return get_object_or_404(Agent, uuid=uuid)

    @extend_schema(
        summary="Retrieve an Agent by UUID",
        parameters=[
            OpenApiParameter(name="uuid", required=True, type=OpenApiTypes.UUID, location=OpenApiParameter.PATH),
        ],
        responses={200: AgentSerializer}
    )
    def get(self, request, uuid):
        item = self.get_object(uuid)
        serializer = AgentSerializer(item)
        return Response(serializer.data)

    @extend_schema(
        summary="Partially update an Agent by UUID",
        request=AgentSerializer,
        responses={200: AgentSerializer}
    )
    def patch(self, request, uuid):
        item = self.get_object(uuid)
        serializer = AgentSerializer(item, data=request.data, partial=True, context={'request': request, 'view': {'kwargs': {'institution_id': item.user.institution_id}}})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @extend_schema(
        summary="Delete an Agent by UUID",
        responses={204: OpenApiResponse(description="Deleted successfully")}
    )
    def delete(self, request, uuid):
        item = self.get_object(uuid)
        item.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)   